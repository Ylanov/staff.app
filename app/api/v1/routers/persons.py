# app/api/v1/routers/persons.py
"""
ИСПРАВЛЕНИЕ: Вся таблица persons в RAM при импорте.

Проблема (старый код):
    existing_records = {p.full_name.lower(): p for p in db.query(Person).all()}
    При 10 000+ записях — вся таблица грузилась в память Python-процесса.

Решение:
    Используем PostgreSQL INSERT ... ON CONFLICT DO UPDATE (upsert) через
    sqlalchemy.dialects.postgresql.insert.
    БД сама обрабатывает конфликты по full_name — Python не держит таблицу в памяти.

    Логика при совпадении full_name:
      - rank, doc_number, department и прочие поля обновляются только если
        в БД значение NULL (COALESCE — не затираем заполненные данные).
      - updated_at обновляется всегда.

    Результат: O(1) память вместо O(N), один SQL-запрос на батч вместо N+1.

    Также исправлен upsert_person_from_slot — теперь тоже использует pg_insert
    вместо SELECT + условного INSERT.
"""

import io
from datetime import date, datetime, timezone
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from sqlalchemy.dialects.postgresql import insert as pg_insert
from pydantic import BaseModel, Field, ConfigDict

from app.db.database import get_db
from app.models.person import Person
from app.models.user import User
from app.api.dependencies import get_current_user, get_current_active_admin

router = APIRouter()

# ─── Константы импорта ────────────────────────────────────────────────────────

CHUNK_SIZE  = 500
MAX_FILE_MB = 10
MAX_ROWS    = 10_000

TEMPLATE_COLS = [
    ("ФИО",             "full_name",      True,  "Иванов Иван Иванович"),
    ("Воинское звание", "rank",           True,  "Майор"),
    ("Номер документа", "doc_number",     False,  "АА 123456"),
    ("Подразделение",   "department",     False, "управление_1"),
    ("Должность",       "position_title", False, "Начальник отдела"),
    ("Дата рождения",   "birth_date",     False, "01.01.1985"),
    ("Телефон",         "phone",          False, "+7 (999) 123-45-67"),
    ("Примечание",      "notes",          False, "Любая заметка"),
]

COL_EXAMPLE = [c[3] for c in TEMPLATE_COLS]


# ─── Схемы ────────────────────────────────────────────────────────────────────

class PersonCreate(BaseModel):
    full_name:      str            = Field(...,  min_length=2, max_length=300, strip_whitespace=True)
    rank:           Optional[str]  = Field(None, max_length=100, strip_whitespace=True)
    doc_number:     Optional[str]  = Field(None, max_length=100, strip_whitespace=True)
    department:     Optional[str]  = Field(None, max_length=100, strip_whitespace=True)
    position_title: Optional[str]  = Field(None, max_length=200, strip_whitespace=True)
    birth_date:     Optional[date] = None
    phone:          Optional[str]  = Field(None, max_length=50,  strip_whitespace=True)
    notes:          Optional[str]  = Field(None, max_length=2000, strip_whitespace=True)


class PersonUpdate(BaseModel):
    full_name:      Optional[str]  = Field(None, min_length=2, max_length=300, strip_whitespace=True)
    rank:           Optional[str]  = Field(None, max_length=100, strip_whitespace=True)
    doc_number:     Optional[str]  = Field(None, max_length=100, strip_whitespace=True)
    department:     Optional[str]  = Field(None, max_length=100, strip_whitespace=True)
    position_title: Optional[str]  = Field(None, max_length=200, strip_whitespace=True)
    birth_date:     Optional[date] = None
    phone:          Optional[str]  = Field(None, max_length=50,  strip_whitespace=True)
    notes:          Optional[str]  = Field(None, max_length=2000, strip_whitespace=True)


class PersonResponse(BaseModel):
    id:             int
    full_name:      str
    rank:           Optional[str]  = None
    doc_number:     Optional[str]  = None
    department:     Optional[str]  = None
    position_title: Optional[str]  = None
    birth_date:     Optional[date] = None
    phone:          Optional[str]  = None
    notes:          Optional[str]  = None

    model_config = ConfigDict(from_attributes=True)


class ImportRowError(BaseModel):
    row:     int
    message: str


class ImportResult(BaseModel):
    message:  str
    added:    int
    updated:  int
    skipped:  int
    errors:   List[ImportRowError] = []


# ─── Вспомогательные функции парсинга ────────────────────────────────────────

def _clean(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() != "none" else None


def _parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _validate_row(row_num: int, fields: dict) -> Optional[str]:
    fn = fields.get("full_name")
    if not fn or len(fn) < 2:
        return f"Строка {row_num}: ФИО обязательно (минимум 2 символа)"
    if not fields.get("rank"):
        return f"Строка {row_num}: Воинское звание обязательно (ФИО: {fn})"
    return None


# ─── Поиск (автодополнение) ───────────────────────────────────────────────────

@router.get(
    "/search",
    response_model=List[PersonResponse],
    summary="Поиск человека по ФИО (для автодополнения)",
)
def search_persons(
        q:     str = Query(..., min_length=2),
        limit: int = Query(10, ge=1, le=50),
        db:    Session = Depends(get_db),
        current_user: User = Depends(get_current_user),
):
    query = db.query(Person).filter(Person.full_name.ilike(f"%{q}%"))
    if current_user.role != "admin":
        query = query.filter(Person.department == current_user.username)
    return query.order_by(Person.full_name).limit(limit).all()


# ─── Скачать шаблон Excel ─────────────────────────────────────────────────────

@router.get(
    "/import/template",
    summary="Скачать шаблон Excel для импорта",
)
def download_import_template(
        current_user: User = Depends(get_current_active_admin),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Шаблон импорта"

    RED_FILL    = PatternFill("solid", fgColor="C0392B")
    GREY_FILL   = PatternFill("solid", fgColor="7F8C8D")
    YELLOW_FILL = PatternFill("solid", fgColor="F9E79F")
    THIN        = Side(style="thin", color="CCCCCC")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    ws.merge_cells("A1:H1")
    note = ws["A1"]
    note.value = (
        "🔴 Красные колонки — ОБЯЗАТЕЛЬНЫЕ.  "
        "⚫ Серые — необязательные.  "
        "Строку 2 (пример) можно удалить.  "
        "Данные начинаются со строки 3."
    )
    note.font      = Font(bold=True, size=10)
    note.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    for col_idx, (header, _, required, _ex) in enumerate(TEMPLATE_COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill      = RED_FILL if required else GREY_FILL
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER
    ws.row_dimensions[2].height = 22

    for col_idx, (_h, _f, _r, example) in enumerate(TEMPLATE_COLS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=example)
        cell.fill      = YELLOW_FILL
        cell.font      = Font(italic=True, color="555555", size=10)
        cell.alignment = Alignment(vertical="center")
        cell.border    = BORDER
    ws.row_dimensions[3].height = 18

    for i, width in enumerate([35, 20, 16, 18, 28, 16, 22, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="persons_template.xlsx"'},
    )


# ─── Массовый импорт из Excel ─────────────────────────────────────────────────

@router.post(
    "/import",
    response_model=ImportResult,
    summary="Массовый импорт из Excel",
)
async def import_persons_from_excel(
        file: UploadFile = File(...),
        db:   Session    = Depends(get_db),
        current_admin: User = Depends(get_current_active_admin),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Формат файла должен быть .xlsx")

    contents = await file.read()

    if len(contents) > MAX_FILE_MB * 1024 * 1024:
        raise HTTPException(
            status_code=400,
            detail=f"Файл слишком большой (максимум {MAX_FILE_MB} МБ)",
        )

    try:
        wb    = openpyxl.load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
        sheet = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось прочитать Excel файл")

    # ── Парсинг и валидация строк ─────────────────────────────────────────────
    errors:       List[ImportRowError] = []
    valid_rows:   List[dict]           = []
    seen_names:   set                  = set()
    skipped_count = 0
    row_num       = 0

    for excel_row in sheet.iter_rows(min_row=2, values_only=True):
        row_num += 1

        if row_num > MAX_ROWS:
            errors.append(ImportRowError(
                row=row_num,
                message=f"Превышен лимит {MAX_ROWS} строк. Остаток проигнорирован."
            ))
            break

        if not any(excel_row):
            skipped_count += 1
            continue

        def get_col(idx: int) -> Optional[str]:
            return _clean(excel_row[idx]) if len(excel_row) > idx else None

        first_val = get_col(0)

        if not first_val:
            skipped_count += 1
            continue

        # Пропускаем строку-заголовок и строку-пример шаблона
        if first_val.lower() in ("фио", "fullname", "full_name") or first_val == COL_EXAMPLE[0]:
            skipped_count += 1
            continue

        fields = {
            "full_name":      first_val,
            "rank":           get_col(1),
            "doc_number":     get_col(2),
            "department":     get_col(3),
            "position_title": get_col(4),
            "birth_date":     _parse_date(excel_row[5]) if len(excel_row) > 5 else None,
            "phone":          get_col(6),
            "notes":          get_col(7),
        }

        err_msg = _validate_row(row_num + 1, fields)
        if err_msg:
            errors.append(ImportRowError(row=row_num + 1, message=err_msg))
            continue

        # Дедупликация внутри самого файла
        key = fields["full_name"].lower()
        if key in seen_names:
            errors.append(ImportRowError(
                row=row_num + 1,
                message=f"Дубль ФИО «{fields['full_name']}» внутри файла — пропущена"
            ))
            skipped_count += 1
            continue
        seen_names.add(key)

        now = datetime.now(timezone.utc)
        valid_rows.append({
            "full_name":      fields["full_name"],
            "rank":           fields["rank"],
            "doc_number":     fields["doc_number"],
            "department":     fields["department"],
            "position_title": fields["position_title"],
            "birth_date":     fields["birth_date"],
            "phone":          fields["phone"],
            "notes":          fields["notes"],
            "created_at":     now,
            "updated_at":     now,
        })

    if not valid_rows:
        return ImportResult(
            message="Нет валидных строк для импорта.",
            added=0, updated=0, skipped=skipped_count, errors=errors,
        )

    # ── PostgreSQL upsert батчами ─────────────────────────────────────────────
    # ИСПРАВЛЕНО: один SQL-запрос на батч вместо загрузки всей таблицы в RAM.
    #
    # INSERT ... ON CONFLICT (full_name) DO UPDATE SET
    #   rank = COALESCE(persons.rank, EXCLUDED.rank),  ← не затираем если уже есть
    #   ...
    #
    # xmax = 0 означает что строка была вставлена (INSERT),
    # xmax != 0 — что обновлена (UPDATE). Так считаем added/updated без SELECT.

    added_count   = 0
    updated_count = 0

    for i in range(0, len(valid_rows), CHUNK_SIZE):
        chunk = valid_rows[i: i + CHUNK_SIZE]

        stmt = pg_insert(Person).values(chunk)
        stmt = stmt.on_conflict_do_update(
            index_elements=["full_name"],
            set_={
                # COALESCE: берём новое значение только если старое NULL
                "rank":           text("COALESCE(persons.rank,           EXCLUDED.rank)"),
                "doc_number":     text("COALESCE(persons.doc_number,     EXCLUDED.doc_number)"),
                "department":     text("COALESCE(persons.department,     EXCLUDED.department)"),
                "position_title": text("COALESCE(persons.position_title, EXCLUDED.position_title)"),
                "birth_date":     text("COALESCE(persons.birth_date,     EXCLUDED.birth_date)"),
                "phone":          text("COALESCE(persons.phone,          EXCLUDED.phone)"),
                "notes":          text("COALESCE(persons.notes,          EXCLUDED.notes)"),
                "updated_at":     stmt.excluded.updated_at,
            },
        )
        # returning xmax: 0 = INSERT, !=0 = UPDATE
        stmt = stmt.returning(
            Person.id,
            text("(xmax = 0)::int AS was_inserted"),
        )

        try:
            result = db.execute(stmt)
            db.commit()
            for row in result.fetchall():
                if row[1]:
                    added_count += 1
                else:
                    updated_count += 1
        except Exception as e:
            db.rollback()
            errors.append(ImportRowError(
                row=0,
                message=f"Ошибка батча {i+1}–{i+len(chunk)}: {str(e)[:120]}"
            ))

    return ImportResult(
        message=(
            f"Импорт завершён. "
            f"Добавлено: {added_count}, обновлено: {updated_count}, "
            f"пропущено: {skipped_count}, ошибок: {len(errors)}."
        ),
        added=added_count,
        updated=updated_count,
        skipped=skipped_count,
        errors=errors,
    )


# ─── CRUD ─────────────────────────────────────────────────────────────────────

@router.get("", response_model=List[PersonResponse], summary="Получить базу людей")
def get_all_persons(
        db:   Session = Depends(get_db),
        current_user: User = Depends(get_current_user),
        skip:  int = Query(0, ge=0),
        limit: int = Query(500, ge=1, le=2000),
        q:     Optional[str] = Query(None),
):
    query = db.query(Person)
    if current_user.role != "admin":
        query = query.filter(Person.department == current_user.username)
    if q:
        query = query.filter(Person.full_name.ilike(f"%{q}%"))
    return query.order_by(Person.full_name).offset(skip).limit(limit).all()


@router.post("", response_model=PersonResponse, status_code=status.HTTP_201_CREATED,
             summary="Добавить человека")
def create_person(
        person_in: PersonCreate,
        db:        Session = Depends(get_db),
        current_user: User = Depends(get_current_user),
):
    department = person_in.department if current_user.role == "admin" else current_user.username

    existing = db.query(Person).filter(Person.full_name.ilike(person_in.full_name)).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT,
                            detail="Человек с таким ФИО уже есть в базе")

    person = Person(
        full_name=      person_in.full_name,
        rank=           person_in.rank           or None,
        doc_number=     person_in.doc_number     or None,
        department=     department,
        position_title= person_in.position_title or None,
        birth_date=     person_in.birth_date,
        phone=          person_in.phone          or None,
        notes=          person_in.notes          or None,
    )
    db.add(person)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT,
                            detail="Человек с таким ФИО уже есть в базе")
    db.refresh(person)
    return person


@router.put("/{person_id}", response_model=PersonResponse, summary="Обновить данные человека")
def update_person(
        person_id: int,
        person_in: PersonUpdate,
        db:        Session = Depends(get_db),
        current_user: User = Depends(get_current_user),
):
    person = db.query(Person).filter(Person.id == person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    if current_user.role != "admin" and person.department != current_user.username:
        raise HTTPException(status_code=403, detail="Нет доступа")

    for field in ("full_name", "rank", "doc_number", "position_title",
                  "birth_date", "phone", "notes"):
        val = getattr(person_in, field, None)
        if val is not None:
            setattr(person, field, val or None)

    if person_in.department is not None and current_user.role == "admin":
        person.department = person_in.department or None

    db.commit()
    db.refresh(person)
    return person


@router.delete("/{person_id}", summary="Удалить человека из базы")
def delete_person(
        person_id: int,
        db:        Session = Depends(get_db),
        current_user: User = Depends(get_current_user),
):
    person = db.query(Person).filter(Person.id == person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail="Человек не найден")
    if current_user.role != "admin" and person.department != current_user.username:
        raise HTTPException(status_code=403, detail="Нет доступа")
    db.delete(person)
    db.commit()
    return {"message": "Удалён из базы"}


# ─── Внутренняя функция: upsert при заполнении слота ─────────────────────────

def upsert_person_from_slot(
        db:         Session,
        full_name:  str,
        rank:       str | None,
        doc_number: str | None,
        department: str | None = None,
) -> None:
    """
    Создаёт или обновляет запись в базе людей при сохранении слота.
    ИСПРАВЛЕНО: использует pg_insert с on_conflict_do_update
    вместо SELECT + условного INSERT (два круговых пути к БД → один).
    """
    if not full_name or not full_name.strip():
        return

    full_name = full_name.strip()
    now = datetime.now(timezone.utc)

    stmt = pg_insert(Person).values(
        full_name=  full_name,
        rank=       rank.strip()       if rank       else None,
        doc_number= doc_number.strip() if doc_number else None,
        department= department.strip() if department else None,
        created_at= now,
        updated_at= now,
    )
    stmt = stmt.on_conflict_do_update(
        index_elements=["full_name"],
        set_={
            "rank":       text("COALESCE(persons.rank,       EXCLUDED.rank)"),
            "doc_number": text("COALESCE(persons.doc_number, EXCLUDED.doc_number)"),
            "department": text("COALESCE(persons.department, EXCLUDED.department)"),
            "updated_at": stmt.excluded.updated_at,
        },
    )

    try:
        db.execute(stmt)
        db.flush()
    except Exception:
        db.rollback()